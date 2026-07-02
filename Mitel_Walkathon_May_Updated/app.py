from flask import (Flask, render_template, request, jsonify,
                   session, redirect, url_for, Response, send_from_directory)
from werkzeug.middleware.proxy_fix import ProxyFix
from datetime import datetime, timedelta, date
import sqlite3, os, csv, io, secrets, string, requests as req_lib
import base64, uuid, logging, threading, subprocess, shutil, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'walkathon-stepup-secret-2025')
app.config['PREFERRED_URL_SCHEME'] = 'https'
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
DB_PATH       = os.path.join(BASE_DIR, 'walkathon.db')
SSL_CERT_PATH = os.environ.get('SSL_CERT_FILE', os.path.join(BASE_DIR, 'cert.pem'))
SSL_KEY_PATH  = os.environ.get('SSL_KEY_FILE',  os.path.join(BASE_DIR, 'key.pem'))

# ─── NETWORK SHARE CONFIG ────────────────────────────────────────────────────
# Hardcoded defaults – override with environment variables if needed
NETWORK_SHARE    = os.environ.get('NETWORK_SHARE',    r'\\10.211.243.115\walkdata')
NETWORK_USER     = os.environ.get('NETWORK_USER',     'mitel')
NETWORK_PASSWORD = os.environ.get('NETWORK_PASSWORD', 'Mitel@123')

# Sub-folders on the share
NETWORK_EXCEL_DIR  = os.path.join(NETWORK_SHARE, 'excel')   # Excel backups
NETWORK_PHOTOS_DIR = os.path.join(NETWORK_SHARE, 'photos')  # Photo uploads

# Linux mount point (used when running on Linux)
LINUX_MOUNT_POINT = os.environ.get('LINUX_MOUNT_POINT', '/mnt/walkdata')

# Local fallback dirs (always used for serving + as safety net)~
LOCAL_BACKUP_DIR  = os.path.join(BASE_DIR, 'backups')
LOCAL_UPLOADS_DIR = os.path.join(BASE_DIR, 'static', 'uploads')

# ─── SECOND BACKUP: OneDrive Desktop folder ───────────────────────────────────
# Syncs automatically via OneDrive client on the server machine
ONEDRIVE_BACKUP_PATH = os.environ.get(
    'ONEDRIVE_BACKUP_PATH',
    r'C:\Users\gowdp\OneDrive - Mitel Networks Corporation\Desktop\backup wakathon'
)
ONEDRIVE_EXCEL_DIR  = os.path.join(ONEDRIVE_BACKUP_PATH, 'excel')
ONEDRIVE_PHOTOS_DIR = os.path.join(ONEDRIVE_BACKUP_PATH, 'photos')

BACKUP_INTERVAL_SECONDS = int(os.environ.get('BACKUP_INTERVAL_SECONDS', '3600'))

FITBIT_AUTH_URL  = "https://www.fitbit.com/oauth2/authorize"
FITBIT_TOKEN_URL = "https://api.fitbit.com/oauth2/token"
FITBIT_STEPS_URL = "https://api.fitbit.com/1/user/-/activities/steps/date/{date}/1d.json"
FITBIT_REDIRECT  = os.environ.get('FITBIT_REDIRECT')

# ─── NETWORK SHARE CONNECTION ────────────────────────────────────────────────

def _is_windows():
    return sys.platform.startswith('win')

def connect_network_share():
    r"""
    Authenticate and connect to the network share.
    Windows : net use  \\server\share /user:USERNAME PASSWORD
    Linux   : mount -t cifs  //server/share /mnt/walkdata -o user=...,pass=...
    Returns True if the share is reachable after the attempt.
    """
    share = NETWORK_SHARE
    if _is_windows():
        try:
            # Disconnect any stale mapping first (ignore errors)
            subprocess.run(['net', 'use', share, '/delete', '/yes'],
                           capture_output=True, timeout=10)
            result = subprocess.run(
                ['net', 'use', share, f'/user:{NETWORK_USER}', NETWORK_PASSWORD, '/persistent:no'],
                capture_output=True, text=True, timeout=15
            )
            if result.returncode == 0:
                logger.info(f'✅ Network share connected: {share}')
                return True
            else:
                logger.warning(f'⚠️  net use failed (rc={result.returncode}): {result.stderr.strip()}')
                return False
        except Exception as e:
            logger.warning(f'⚠️  connect_network_share (Windows) exception: {e}')
            return False
    else:
        # Linux: mount via cifs
        try:
            # Convert Windows UNC to Linux host/path  \\10.x.x.x\share → //10.x.x.x/share
            unc_linux = share.replace('\\', '/').lstrip('/')
            cifs_path = '//' + unc_linux
            mp = LINUX_MOUNT_POINT
            os.makedirs(mp, exist_ok=True)
            # Check if already mounted
            with open('/proc/mounts') as f:
                if mp in f.read():
                    logger.info(f'✅ Network share already mounted at {mp}')
                    return True
            result = subprocess.run(
                ['mount', '-t', 'cifs', cifs_path, mp,
                 '-o', f'username={NETWORK_USER},password={NETWORK_PASSWORD},vers=3.0,iocharset=utf8'],
                capture_output=True, text=True, timeout=20
            )
            if result.returncode == 0:
                logger.info(f'✅ Network share mounted at {mp}')
                return True
            else:
                logger.warning(f'⚠️  cifs mount failed: {result.stderr.strip()}')
                return False
        except Exception as e:
            logger.warning(f'⚠️  connect_network_share (Linux) exception: {e}')
            return False

def _network_excel_dir():
    """Return the effective Excel backup directory (network or local fallback)."""
    if _is_windows():
        return NETWORK_EXCEL_DIR
    else:
        mp = LINUX_MOUNT_POINT
        try:
            if os.path.ismount(mp):
                return os.path.join(mp, 'excel')
        except Exception:
            pass
        return LOCAL_BACKUP_DIR

def _network_photos_dir():
    """Return the effective photos directory on the network share."""
    if _is_windows():
        return NETWORK_PHOTOS_DIR
    else:
        mp = LINUX_MOUNT_POINT
        try:
            if os.path.ismount(mp):
                return os.path.join(mp, 'photos')
        except Exception:
            pass
        return None   # no network photos dir available — local only

def _ensure_network_dirs():
    """Create sub-folders on the share AND OneDrive folder if they don't exist yet."""
    all_dirs = [_network_excel_dir(), _network_photos_dir(),
                ONEDRIVE_EXCEL_DIR, ONEDRIVE_PHOTOS_DIR, LOCAL_BACKUP_DIR]
    for d in all_dirs:
        if d:
            try:
                os.makedirs(d, exist_ok=True)
            except Exception as e:
                logger.warning(f'Could not create dir {d}: {e}')

# ─── PHOTO HELPERS ───────────────────────────────────────────────────────────

def _copy_photo_to_destinations_bg(filename: str, data: bytes):
    """Background thread: copy photo to network share + OneDrive (non-blocking)."""
    for dest_dir, label in [(_network_photos_dir(), 'network share'), (ONEDRIVE_PHOTOS_DIR, 'OneDrive')]:
        if not dest_dir:
            continue
        try:
            os.makedirs(dest_dir, exist_ok=True)
            dest_path = os.path.join(dest_dir, filename)
            with open(dest_path, 'wb') as f:
                f.write(data)
            logger.info(f'📸 Photo backed up to {label}: {dest_path}')
        except Exception as e:
            logger.warning(f'Could not copy photo to {label}: {e}')

def save_photo_file(filename: str, data: bytes):
    """
    Save photo bytes to:
      1. Local  static/uploads/<filename>  (always – primary, used for serving)
      2. Network share photos/<filename>   (backup – background thread, non-blocking)
      3. OneDrive photos/<filename>        (backup – background thread, non-blocking)
    """
    # Always save locally first so the upload never fails due to network issues
    os.makedirs(LOCAL_UPLOADS_DIR, exist_ok=True)
    local_path = os.path.join(LOCAL_UPLOADS_DIR, filename)
    with open(local_path, 'wb') as f:
        f.write(data)

    # Copy to remote destinations in a background thread – never blocks the user
    t = threading.Thread(target=_copy_photo_to_destinations_bg, args=(filename, data), daemon=True)
    t.start()

def delete_photo_file(filename: str):
    """Delete photo from local uploads, network share, and OneDrive (best-effort)."""
    for d in [LOCAL_UPLOADS_DIR, _network_photos_dir(), ONEDRIVE_PHOTOS_DIR]:
        if not d:
            continue
        path = os.path.join(d, filename)
        if os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass

def restore_photo_from_network(filename: str) -> bool:
    """
    If a photo is missing locally, try to restore it from the network share or OneDrive.
    Returns True if restored.
    """
    local_path = os.path.join(LOCAL_UPLOADS_DIR, filename)
    if os.path.exists(local_path):
        return True  # already there

    # Try network share first, then OneDrive
    for src_dir, label in [(_network_photos_dir(), 'network share'), (ONEDRIVE_PHOTOS_DIR, 'OneDrive')]:
        if not src_dir:
            continue
        src_path = os.path.join(src_dir, filename)
        if not os.path.exists(src_path):
            continue
        try:
            os.makedirs(LOCAL_UPLOADS_DIR, exist_ok=True)
            shutil.copy2(src_path, local_path)
            logger.info(f'🔄 Photo restored from {label}: {filename}')
            return True
        except Exception as e:
            logger.warning(f'Could not restore photo {filename} from {label}: {e}')

    return False

# ─── DB ──────────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = NORMAL")
    return conn

def init_db():
    conn = get_db(); c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS teams (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL,
        invite_code TEXT UNIQUE NOT NULL, created_by INTEGER,
        created_at TEXT DEFAULT (datetime('now','localtime')))''')
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
        email TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL,
        role TEXT DEFAULT 'user', team_id INTEGER REFERENCES teams(id),
        weight_kg REAL, fitbit_access_token TEXT, fitbit_refresh_token TEXT,
        fitbit_user_id TEXT, country TEXT, country_code TEXT, country_flag TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime')))''')
    c.execute('''CREATE TABLE IF NOT EXISTS steps (
        id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL REFERENCES users(id),
        steps INTEGER NOT NULL, date TEXT NOT NULL, note TEXT,
        source TEXT DEFAULT 'manual', session_id TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        updated_at TEXT DEFAULT (datetime('now','localtime')))''')
    steps_cols = {r[1] for r in c.execute("PRAGMA table_info(steps)").fetchall()}
    if 'session_id' not in steps_cols:
        c.execute("ALTER TABLE steps ADD COLUMN session_id TEXT")
    c.execute("CREATE INDEX IF NOT EXISTS idx_steps_user_date ON steps(user_id, date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_steps_user_session ON steps(user_id, date, session_id)")
    users_cols = {r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()}
    for col in ('country','country_code','country_flag'):
        if col not in users_cols:
            c.execute(f"ALTER TABLE users ADD COLUMN {col} TEXT")
    c.execute("""UPDATE users SET country_flag=
        CASE WHEN country_code IS NOT NULL AND LENGTH(country_code)=2 THEN
            CHAR(127397+UNICODE(SUBSTR(UPPER(country_code),1,1)))||CHAR(127397+UNICODE(SUBSTR(UPPER(country_code),2,1)))
        ELSE country_flag END WHERE country_flag IS NULL AND country_code IS NOT NULL""")
    c.execute('''CREATE TABLE IF NOT EXISTS config (
        id INTEGER PRIMARY KEY CHECK(id=1),
        event_name TEXT DEFAULT 'Mitel Walk-a-Thon 2026',
        daily_goal INTEGER DEFAULT 10000, fitbit_client_id TEXT DEFAULT '',
        fitbit_client_secret TEXT DEFAULT '', start_date TEXT DEFAULT '', end_date TEXT DEFAULT '')''')
    c.execute('''CREATE TABLE IF NOT EXISTS notices (
        id INTEGER PRIMARY KEY AUTOINCREMENT, message TEXT NOT NULL,
        created_by TEXT DEFAULT 'Admin', created_at TEXT DEFAULT (datetime('now','localtime')),
        is_active INTEGER DEFAULT 1)''')
    c.execute('''CREATE TABLE IF NOT EXISTS gallery (
        id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL REFERENCES users(id),
        user_name TEXT NOT NULL, thought TEXT, image_filename TEXT NOT NULL,
        image_type TEXT DEFAULT 'image/jpeg',
        created_at TEXT DEFAULT (datetime('now','localtime')),
        updated_at TEXT DEFAULT (datetime('now','localtime')))''')
    gallery_cols = {r[1] for r in c.execute("PRAGMA table_info(gallery)").fetchall()}
    if 'updated_at' not in gallery_cols:
        c.execute("ALTER TABLE gallery ADD COLUMN updated_at TEXT")
        c.execute("UPDATE gallery SET updated_at=COALESCE(updated_at,created_at)")
    c.execute("INSERT OR IGNORE INTO config (id) VALUES (1)")
    ADMIN_PASSWORD = 'Flask@Mitel#Walkathon26!'
    admin_row = conn.execute("SELECT id,password_hash FROM users WHERE role='admin' AND email='admin@walkathon.com'").fetchone()
    if not admin_row:
        conn.execute("INSERT INTO users (name,email,password_hash,role) VALUES (?,?,?,?)",
                     ('Admin','admin@walkathon.com',ADMIN_PASSWORD,'admin'))
    else:
        stored = admin_row['password_hash'] or ''
        if stored.startswith('pbkdf2:') or stored.startswith('scrypt:'):
            conn.execute("UPDATE users SET password_hash=? WHERE id=?", (ADMIN_PASSWORD, admin_row['id']))
    conn.commit(); conn.close()

# ─── GENERAL HELPERS ─────────────────────────────────────────────────────────

def _dash(v):
    return '-' if (v is None or v == '') else v

def steps_to_km(s):  return round((s or 0)/1300, 2)
def steps_to_cal(s, w):
    if not w: return None
    rate = 0.04 if w<60 else (0.05 if w<=85 else 0.06)
    return round((s or 0)*rate)

def gen_invite():
    return ''.join(secrets.choice(string.ascii_uppercase+string.digits) for _ in range(6))

def code_to_flag(code):
    if not code or len(code)!=2 or not code.isalpha(): return ''
    code=code.upper()
    return chr(127397+ord(code[0]))+chr(127397+ord(code[1]))

def normalize_country(country, country_code, country_flag):
    country=(country or '').strip(); country_code=(country_code or '').strip().upper(); country_flag=(country_flag or '').strip()
    if country and len(country)>=3 and country[:2].isalpha() and country[2]==' ':
        inferred=country[:2].upper(); remainder=country[3:].strip()
        if remainder:
            country_code=country_code or inferred; country=remainder
            if not country_flag: country_flag=code_to_flag(inferred)
    if country_flag and len(country_flag)==2 and country_flag.isalpha(): country_flag=code_to_flag(country_flag)
    if not country_flag and country_code: country_flag=code_to_flag(country_code)
    display=f"{country_flag} {country}" if (country and country_flag) else (country or '—')
    return country, country_code, country_flag, display

def today_str(): return date.today().isoformat()

def calc_streak(user_id):
    conn=get_db(); rows=conn.execute("SELECT DISTINCT date FROM steps WHERE user_id=? ORDER BY date DESC",(user_id,)).fetchall(); conn.close()
    if not rows: return 0
    dates={r['date'] for r in rows}; d=date.today(); streak=0
    while str(d) in dates: streak+=1; d-=timedelta(days=1)
    if streak==0:
        d=date.today()-timedelta(days=1)
        while str(d) in dates: streak+=1; d-=timedelta(days=1)
    return streak

def get_badges(total, streak):
    b=[]
    if total>=10000:  b.append({'name':'10K Champion','icon':'🏆','color':'#FFD700'})
    if total>=50000:  b.append({'name':'Step Legend','icon':'⭐','color':'#FF6B6B'})
    if total>=100000: b.append({'name':'Ultra Walker','icon':'🔥','color':'#FF4500'})
    if streak>=3:     b.append({'name':'Consistency King','icon':'👑','color':'#9B59B6'})
    if streak>=7:     b.append({'name':'7-Day Warrior','icon':'⚡','color':'#3498DB'})
    if steps_to_km(total)>=10: b.append({'name':'10km Crusher','icon':'🎯','color':'#2ECC71'})
    return b

def motivation(today_s, goal):
    pct=(today_s/goal*100) if goal else 0
    if today_s==0: return {'msg':'Every step counts! Start today 🌟','level':'low'}
    if pct<25:     return {'msg':'Great start! Keep moving 🚶','level':'low'}
    if pct<50:     return {'msg':'Building momentum! Push forward 💪','level':'medium'}
    if pct<75:     return {'msg':'More than halfway! Keep it up 🔥','level':'medium'}
    if pct<100:    return {'msg':'Almost there! One final push 🎯','level':'high'}
    return {'msg':'GOAL ACHIEVED! Legend! 🏆🎉','level':'champion'}

def current_user():
    uid=session.get('user_id')
    if not uid: return None
    conn=get_db(); u=conn.execute("SELECT u.*,t.name as team_name,t.invite_code FROM users u LEFT JOIN teams t ON t.id=u.team_id WHERE u.id=?",(uid,)).fetchone(); conn.close()
    return u

def is_gallery_manager(row,uid=None,role=None):
    uid=session.get('user_id') if uid is None else uid; role=session.get('role') if role is None else role
    return role=='admin' or row['user_id']==uid

def serialize_gallery_row(row,uid=None,role=None):
    return {'id':row['id'],'user_id':row['user_id'],'user_name':row['user_name'],
            'thought':row['thought'] or '',
            'image_url':url_for('serve_upload',filename=row['image_filename']),
            'image_filename':row['image_filename'],'image_type':row['image_type'],
            'created_at':row['created_at'],'updated_at':row['updated_at'] or row['created_at'],
            'is_owner':row['user_id']==uid,'can_manage':is_gallery_manager(row,uid=uid,role=role)}

def validate_gallery_payload(data,require_image=False):
    image_data=(data.get('image_data') or '').strip(); image_type=(data.get('image_type') or 'image/jpeg').strip().lower()
    thought=(data.get('thought') or '').strip(); allowed={'image/jpeg','image/jpg','image/png','image/webp','image/gif'}
    if require_image and not image_data: return None,('No image provided',400)
    if thought and len(thought)>300: return None,('Thought must be 300 characters or less',400)
    if image_data:
        if image_type not in allowed: return None,('Unsupported image type',400)
        try: decoded=base64.b64decode(image_data)
        except: return None,('Invalid image data',400)
        if len(decoded)>5_000_000: return None,('Image file size exceeds 5MB',400)
    return {'image_data':image_data,'image_type':image_type,'thought':thought or None},None

def env_flag(name,default=False):
    v=os.environ.get(name)
    return default if v is None else v.strip().lower() in {'1','true','yes','on'}

def get_ssl_context():
    cert=os.path.exists(SSL_CERT_PATH); key=os.path.exists(SSL_KEY_PATH)
    if not env_flag('USE_HTTPS', cert and key): return None
    return (SSL_CERT_PATH,SSL_KEY_PATH) if (cert and key) else None

def login_required(f):
    from functools import wraps
    @wraps(f)
    def d(*a,**k):
        if not session.get('user_id'): return redirect(url_for('login_page'))
        return f(*a,**k)
    return d

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def d(*a,**k):
        if session.get('role')!='admin': return redirect(url_for('login_page'))
        return f(*a,**k)
    return d

# ─── EXCEL WORKBOOK BUILDER ───────────────────────────────────────────────────

def _style_header(ws, row_num, col_count, hex_color='1F6AA5'):
    fill=PatternFill("solid",fgColor=hex_color); font=Font(bold=True,color='FFFFFF',size=11)
    align=Alignment(horizontal='center',vertical='center',wrap_text=True)
    thin=Side(style='thin',color='AAAAAA'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    for col in range(1,col_count+1):
        cell=ws.cell(row=row_num,column=col)
        cell.fill=fill; cell.font=font; cell.alignment=align; cell.border=border

def _auto_width(ws):
    for col in ws.columns:
        ml=max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(ml+4,50)

def build_excel_workbook():
    conn=get_db(); wb=openpyxl.Workbook(); wb.remove(wb.active)

    # Users (human-readable)
    ws=wb.create_sheet('Users')
    hdrs=['ID','Name','Email','Password','Role','Team','Weight (kg)','Country','Country Code','Total Steps','Total KM','Calories','Joined At']
    ws.append(hdrs); _style_header(ws,1,len(hdrs),'1F6AA5')
    for u in conn.execute('''SELECT u.id,u.name,u.email,u.password_hash,u.role,t.name as team_name,u.weight_kg,u.country,u.country_code,u.created_at,COALESCE(SUM(s.steps),0) as total_steps
        FROM users u LEFT JOIN teams t ON t.id=u.team_id LEFT JOIN steps s ON s.user_id=u.id WHERE u.role!='admin' GROUP BY u.id ORDER BY total_steps DESC''').fetchall():
        tot=u['total_steps']; cal=steps_to_cal(tot,u['weight_kg'])
        ws.append([u['id'],u['name'],u['email'],u['password_hash'],u['role'],_dash(u['team_name']),_dash(u['weight_kg']),_dash(u['country']),_dash(u['country_code']),tot,steps_to_km(tot),_dash(cal),u['created_at'] or '-'])
    _auto_width(ws)

    # Teams (human-readable)
    ws2=wb.create_sheet('Teams')
    hdrs2=['ID','Team Name','Invite Code','Members','Total Steps','Created At']
    ws2.append(hdrs2); _style_header(ws2,1,len(hdrs2),'2E7D32')
    for t in conn.execute('''SELECT t.id,t.name,t.invite_code,t.created_at,COUNT(DISTINCT u.id) as members,COALESCE(SUM(s.steps),0) as total_steps
        FROM teams t LEFT JOIN users u ON u.team_id=t.id LEFT JOIN steps s ON s.user_id=u.id GROUP BY t.id ORDER BY total_steps DESC''').fetchall():
        ws2.append([t['id'],t['name'],t['invite_code'],t['members'],t['total_steps'],t['created_at'] or '-'])
    _auto_width(ws2)

    # Steps (human-readable)
    ws3=wb.create_sheet('Steps')
    hdrs3=['Entry ID','User ID','User Name','Team','Steps','KM','Calories','Date','Note','Source','Logged At']
    ws3.append(hdrs3); _style_header(ws3,1,len(hdrs3),'6A1B9A')
    for s in conn.execute('''SELECT s.id,s.user_id,u.name as user_name,t.name as team_name,u.weight_kg,s.steps,s.date,s.note,s.source,s.created_at
        FROM steps s JOIN users u ON u.id=s.user_id LEFT JOIN teams t ON t.id=u.team_id ORDER BY s.date DESC,s.created_at DESC''').fetchall():
        cal=steps_to_cal(s['steps'],s['weight_kg'])
        ws3.append([s['id'],s['user_id'],s['user_name'],_dash(s['team_name']),s['steps'],steps_to_km(s['steps']),_dash(cal),s['date'],_dash(s['note']),s['source'] or '-',s['created_at'] or '-'])
    _auto_width(ws3)

    # Photos index (for reference)
    ws4=wb.create_sheet('Photos')
    hdrs4=['ID','User ID','User Name','Filename','Thought','Uploaded At']
    ws4.append(hdrs4); _style_header(ws4,1,len(hdrs4),'00796B')
    for g in conn.execute("SELECT id,user_id,user_name,image_filename,thought,created_at FROM gallery ORDER BY created_at DESC").fetchall():
        ws4.append([g['id'],g['user_id'],g['user_name'],g['image_filename'],_dash(g['thought']),g['created_at'] or '-'])
    _auto_width(ws4)

    # ── Raw restore sheets ────────────────────────────────────────────────────
    ruw=wb.create_sheet('_RestoreData_Users')
    ruh=['id','name','email','password_hash','role','team_id','weight_kg','country','country_code','country_flag','created_at']
    ruw.append(ruh); _style_header(ruw,1,len(ruh),'455A64')
    for u in conn.execute("SELECT id,name,email,password_hash,role,team_id,weight_kg,country,country_code,country_flag,created_at FROM users WHERE role!='admin'").fetchall():
        ruw.append([u['id'],u['name'],u['email'],u['password_hash'],u['role'],u['team_id'],u['weight_kg'],u['country'],u['country_code'],u['country_flag'],u['created_at']])
    _auto_width(ruw)

    rtw=wb.create_sheet('_RestoreData_Teams')
    rth=['id','name','invite_code','created_by','created_at']
    rtw.append(rth); _style_header(rtw,1,len(rth),'455A64')
    for t in conn.execute("SELECT id,name,invite_code,created_by,created_at FROM teams").fetchall():
        rtw.append([t['id'],t['name'],t['invite_code'],t['created_by'],t['created_at']])
    _auto_width(rtw)

    rsw=wb.create_sheet('_RestoreData_Steps')
    rsh=['id','user_id','steps','date','note','source','session_id','created_at','updated_at']
    rsw.append(rsh); _style_header(rsw,1,len(rsh),'455A64')
    for s in conn.execute("SELECT id,user_id,steps,date,note,source,session_id,created_at,updated_at FROM steps").fetchall():
        rsw.append([s['id'],s['user_id'],s['steps'],s['date'],s['note'],s['source'],s['session_id'],s['created_at'],s['updated_at']])
    _auto_width(rsw)

    conn.close()
    return wb

# ─── NETWORK BACKUP ───────────────────────────────────────────────────────────

def _save_excel_to_dir(wb, folder, fname, label):
    """Save workbook to a folder. Returns path on success, None on failure."""
    try:
        os.makedirs(folder, exist_ok=True)
        path = os.path.join(folder, fname)
        wb.save(path)
        logger.info(f'✅ Excel backup → {label}: {path}')
        return path
    except Exception as e:
        logger.warning(f'⚠️  Excel backup FAILED → {label}: {e}')
        return None

def sync_photos_to_dir(dest_dir, label):
    """Copy any photos from static/uploads/ that are missing in dest_dir."""
    if not os.path.isdir(LOCAL_UPLOADS_DIR):
        return 0
    try:
        os.makedirs(dest_dir, exist_ok=True)
    except Exception as e:
        logger.warning(f'Cannot create {label} photos dir: {e}')
        return 0
    files = [f for f in os.listdir(LOCAL_UPLOADS_DIR)
             if os.path.isfile(os.path.join(LOCAL_UPLOADS_DIR, f))]
    synced = 0
    for fname in files:
        dst = os.path.join(dest_dir, fname)
        if os.path.exists(dst):
            continue
        try:
            shutil.copy2(os.path.join(LOCAL_UPLOADS_DIR, fname), dst)
            synced += 1
        except Exception as e:
            logger.warning(f'Could not sync photo {fname} to {label}: {e}')
    if synced:
        logger.info(f'📸 {synced} photo(s) synced → {label}')
    return synced

def sync_all_photos_to_backup_dirs():
    """Sync all local photos to every backup destination."""
    # OneDrive (primary desktop backup – always attempted)
    sync_photos_to_dir(ONEDRIVE_PHOTOS_DIR, 'OneDrive Desktop')
    # Network share (secondary – attempt only if accessible)
    nd = _network_photos_dir()
    if nd:
        sync_photos_to_dir(nd, 'Network share')

def export_excel_to_network():
    """
    Hourly backup job:
      1. Save Excel to OneDrive Desktop folder  (PRIMARY – always attempted)
      2. Save Excel to local backups/ folder    (local safety copy)
      3. Save Excel to network share            (secondary – attempted if reachable)
      4. Sync all photos to OneDrive + network share
    Each destination is independent – failure in one never affects the others.
    """
    wb   = build_excel_workbook()
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    fname = f'walkathon_backup_{ts}.xlsx'
    saved_paths = []

    # ── 1. OneDrive Desktop (PRIMARY) ─────────────────────────────────────────
    p = _save_excel_to_dir(wb, ONEDRIVE_EXCEL_DIR, fname, 'OneDrive Desktop')
    if p: saved_paths.append(p)

    # ── 2. Local backups/ folder ──────────────────────────────────────────────
    p = _save_excel_to_dir(wb, LOCAL_BACKUP_DIR, fname, 'Local backups/')
    if p: saved_paths.append(p)

    # ── 3. Network share ──────────────────────────────────────────────────────
    net_dir = _network_excel_dir()
    if net_dir:
        p = _save_excel_to_dir(wb, net_dir, fname, 'Network share')
        if p: saved_paths.append(p)

    # ── Cleanup old files (keep last 48 = 48 hrs) ─────────────────────────────
    for folder in [ONEDRIVE_EXCEL_DIR, LOCAL_BACKUP_DIR, net_dir]:
        if folder and os.path.isdir(folder):
            _cleanup_old_backups(folder, keep=48)

    # ── Sync photos to OneDrive + network share ───────────────────────────────
    try:
        sync_all_photos_to_backup_dirs()
    except Exception as e:
        logger.warning(f'Photo sync error: {e}')

    logger.info(f'🕐 Hourly backup complete. Saved to {len(saved_paths)} location(s).')
    return saved_paths[0] if saved_paths else None

def _cleanup_old_backups(folder, keep=48):
    try:
        files = sorted([f for f in os.listdir(folder) if f.startswith('walkathon_backup_') and f.endswith('.xlsx')], reverse=True)
        for old in files[keep:]:
            try: os.remove(os.path.join(folder, old))
            except: pass
    except Exception:
        pass

_backup_timer = None

def _schedule_next():
    global _backup_timer
    _backup_timer = threading.Timer(BACKUP_INTERVAL_SECONDS, _run_backup)
    _backup_timer.daemon = True
    _backup_timer.start()

def _run_backup():
    export_excel_to_network()
    _schedule_next()

def start_backup_scheduler():
    logger.info(f'🕐 Auto-backup every {BACKUP_INTERVAL_SECONDS}s')
    logger.info(f'   Local  → {LOCAL_BACKUP_DIR}')
    logger.info(f'   Network→ {NETWORK_SHARE}\\excel')
    _schedule_next()

# ─── RESTORE FROM EXCEL ───────────────────────────────────────────────────────

def _rows_as_dicts(ws):
    rows=list(ws.iter_rows(values_only=True))
    if not rows: return
    headers=[str(h).strip() if h is not None else '' for h in rows[0]]
    for row in rows[1:]:
        if all(v is None for v in row): continue
        yield dict(zip(headers,row))

def _v(row,key,default=None):
    val=row.get(key,default)
    return default if (val=='-' or val=='' or val is None) else val

def restore_from_excel(file_stream):
    try: wb=openpyxl.load_workbook(file_stream,read_only=True,data_only=True)
    except Exception as e: return 0,f'Cannot read Excel file: {e}'
    sheets=wb.sheetnames; conn=get_db(); restored=0
    try:
        use_raw=('_RestoreData_Teams' in sheets and '_RestoreData_Users' in sheets)
        if use_raw:
            restored+=_restore_raw_teams(wb['_RestoreData_Teams'],conn)
            restored+=_restore_raw_users(wb['_RestoreData_Users'],conn)
            if '_RestoreData_Steps' in sheets:
                restored+=_restore_raw_steps(wb['_RestoreData_Steps'],conn)
        else:
            if 'Teams' in sheets: restored+=_restore_teams_fb(wb['Teams'],conn)
            if 'Users' in sheets: restored+=_restore_users_fb(wb['Users'],conn)
        conn.commit()
    except Exception as e:
        conn.rollback(); conn.close(); return 0,f'Restore failed: {e}'
    finally: conn.close()
    return restored,None

def _restore_raw_teams(ws,conn):
    count=0
    for row in _rows_as_dicts(ws):
        name=_v(row,'name'); inv=_v(row,'invite_code'); tid=_v(row,'id')
        ca=_v(row,'created_at',datetime.now().strftime('%Y-%m-%d %H:%M:%S')); cb=_v(row,'created_by')
        if not name or not inv: continue
        if conn.execute("SELECT id FROM teams WHERE name=? OR invite_code=?",(name,inv)).fetchone(): continue
        conn.execute("INSERT OR IGNORE INTO teams (id,name,invite_code,created_by,created_at) VALUES (?,?,?,?,?)",(tid,name,inv,cb,ca)); count+=1
    return count

def _restore_raw_users(ws,conn):
    count=0
    for row in _rows_as_dicts(ws):
        uid=_v(row,'id'); name=_v(row,'name'); email=_v(row,'email')
        pwd=_v(row,'password_hash') or ''; role=_v(row,'role','user')
        tid=_v(row,'team_id'); wkg=_v(row,'weight_kg')
        cnt=_v(row,'country'); cc=_v(row,'country_code'); cf=_v(row,'country_flag')
        ca=_v(row,'created_at',datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        if not email or not name: continue
        if conn.execute("SELECT id FROM users WHERE email=?",(email,)).fetchone(): continue
        if tid and not conn.execute("SELECT id FROM teams WHERE id=?",(tid,)).fetchone(): tid=None
        conn.execute("INSERT OR IGNORE INTO users (id,name,email,password_hash,role,team_id,weight_kg,country,country_code,country_flag,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                     (uid,name,email,pwd,role,tid,wkg,cnt,cc,cf,ca)); count+=1
    return count

def _restore_raw_steps(ws,conn):
    count=0
    for row in _rows_as_dicts(ws):
        sid=_v(row,'id'); uid=_v(row,'user_id'); st=_v(row,'steps'); dt=_v(row,'date')
        note=_v(row,'note'); src=_v(row,'source','manual'); sess=_v(row,'session_id')
        ca=_v(row,'created_at',datetime.now().strftime('%Y-%m-%d %H:%M:%S')); ua=_v(row,'updated_at',ca)
        if not uid or not st or not dt: continue
        if not conn.execute("SELECT id FROM users WHERE id=?",(uid,)).fetchone(): continue
        if conn.execute("SELECT id FROM steps WHERE id=?",(sid,)).fetchone(): continue
        conn.execute("INSERT OR IGNORE INTO steps (id,user_id,steps,date,note,source,session_id,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?)",
                     (sid,uid,int(st),str(dt),note,src,sess,ca,ua)); count+=1
    return count

def _restore_teams_fb(ws,conn):
    count=0
    for row in _rows_as_dicts(ws):
        name=_v(row,'Team Name'); inv=_v(row,'Invite Code') or gen_invite()
        if not name: continue
        if conn.execute("SELECT id FROM teams WHERE name=?",(name,)).fetchone(): continue
        conn.execute("INSERT INTO teams (name,invite_code) VALUES (?,?)",(name,inv)); count+=1
    return count

def _restore_users_fb(ws,conn):
    count=0
    for row in _rows_as_dicts(ws):
        name=_v(row,'Name'); email=_v(row,'Email'); pwd=_v(row,'Password') or 'ChangeMeNow123!'
        team_nm=_v(row,'Team'); wkg=_v(row,'Weight (kg)'); cnt=_v(row,'Country'); cc=_v(row,'Country Code')
        if not email or not name: continue
        if conn.execute("SELECT id FROM users WHERE email=?",(email,)).fetchone(): continue
        tid=None
        if team_nm:
            t=conn.execute("SELECT id FROM teams WHERE name=?",(team_nm,)).fetchone()
            if t: tid=t['id']
        conn.execute("INSERT INTO users (name,email,password_hash,team_id,weight_kg,country,country_code,country_flag) VALUES (?,?,?,?,?,?,?,?)",
                     (name,email,pwd,tid,float(wkg) if wkg else None,cnt,cc,code_to_flag(cc) if cc else None)); count+=1
    return count

# ─── ERROR HANDLERS ──────────────────────────────────────────────────────────

@app.errorhandler(404)
def not_found(e): return jsonify({'error':'Not found'}),404
@app.errorhandler(500)
def internal_error(e): return jsonify({'error':'Internal server error'}),500
@app.errorhandler(Exception)
def handle_exception(e): logger.error(f"Unhandled: {e}"); return jsonify({'error':'Unexpected error'}),500

@app.route('/health')
def health_check():
    try:
        conn=get_db(); conn.execute("SELECT 1").fetchone(); conn.close()
        net_ok = os.path.isdir(_network_excel_dir()) if _network_excel_dir() else False
        return jsonify({'status':'healthy','timestamp':datetime.now().isoformat(),'network_share':net_ok})
    except Exception as e:
        return jsonify({'status':'unhealthy','error':str(e)}),500

# ─── AUTH ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if session.get('user_id'):
        return redirect(url_for('admin_dash') if session.get('role')=='admin' else url_for('dashboard'))
    return redirect(url_for('login_page'))

@app.route('/login',methods=['GET','POST'])
def login_page():
    if session.get('user_id'): return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/api/login',methods=['POST'])
def do_login():
    d=request.json; email=(d.get('email') or '').strip().lower(); password=d.get('password',''); is_admin=d.get('admin',False)
    conn=get_db(); user=conn.execute("SELECT * FROM users WHERE email=?",(email,)).fetchone(); conn.close()
    if not user: return jsonify({'success':False,'message':'Invalid email or password'}),401
    stored=user['password_hash'] or ''
    if stored.startswith('pbkdf2:') or stored.startswith('scrypt:'):
        try:
            from werkzeug.security import check_password_hash as _chk
            ok=_chk(stored,password)
        except: ok=False
    else:
        ok=(stored==password)
    if not ok: return jsonify({'success':False,'message':'Invalid email or password'}),401
    if is_admin and user['role']!='admin': return jsonify({'success':False,'message':'No admin access'}),403
    session['user_id']=user['id']; session['user_name']=user['name']; session['role']=user['role']
    return jsonify({'success':True,'redirect':url_for('admin_dash') if user['role']=='admin' else url_for('dashboard')})

@app.route('/api/logout',methods=['POST'])
def do_logout(): session.clear(); return jsonify({'success':True})

@app.route('/register')
def register_page():
    if session.get('user_id'): return redirect(url_for('index'))
    return render_template('register.html')

@app.route('/api/register',methods=['POST'])
def do_register():
    d=request.json; name=(d.get('name') or '').strip(); email=(d.get('email') or '').strip().lower()
    password=d.get('password',''); weight=d.get('weight_kg'); country=(d.get('country') or '').strip()
    country_code=(d.get('country_code') or '').strip().upper(); country_flag=code_to_flag(country_code)
    team_mode=d.get('team_mode','new'); team_name=(d.get('team_name') or '').strip()
    invite=(d.get('invite_code') or '').strip().upper()
    if not all([name,email,password,country]): return jsonify({'success':False,'message':'Name, email, password and country are required'}),400
    if '@' not in email: return jsonify({'success':False,'message':'Invalid email address'}),400
    if len(password)<9: return jsonify({'success':False,'message':'Password must be at least 9 characters'}),400
    weight_kg=float(weight) if weight and str(weight).strip() else None
    conn=get_db()
    try:
        if conn.execute("SELECT id FROM users WHERE email=?",(email,)).fetchone():
            return jsonify({'success':False,'message':'Email already registered'}),400
        team_id=None
        if team_mode=='new' and team_name:
            if conn.execute("SELECT id FROM teams WHERE LOWER(name)=LOWER(?)",(team_name,)).fetchone():
                return jsonify({'success':False,'message':'Team name already exists. Join it instead.'}),400
            cur=conn.execute("INSERT INTO teams (name,invite_code) VALUES (?,?)",(team_name,gen_invite()))
            team_id=cur.lastrowid
        elif team_mode=='join' and invite:
            t=conn.execute("SELECT id FROM teams WHERE invite_code=?",(invite,)).fetchone()
            if not t: return jsonify({'success':False,'message':'Invalid invite code'}),400
            team_id=t['id']
        cur=conn.execute("INSERT INTO users (name,email,password_hash,team_id,weight_kg,country,country_code,country_flag) VALUES (?,?,?,?,?,?,?,?)",
                         (name,email,password,team_id,weight_kg,country,country_code,country_flag))
        conn.commit(); uid=cur.lastrowid
        session['user_id']=uid; session['user_name']=name; session['role']='user'
        return jsonify({'success':True,'redirect':url_for('dashboard')})
    except Exception as e:
        return jsonify({'success':False,'message':str(e)}),500
    finally: conn.close()

@app.route('/api/teams/search')
def search_teams():
    q=request.args.get('q','').strip(); conn=get_db()
    rows=conn.execute("SELECT id,name,invite_code,(SELECT COUNT(*) FROM users WHERE team_id=teams.id) as members FROM teams WHERE LOWER(name) LIKE LOWER(?) LIMIT 10",(f'%{q}%',)).fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

# ─── GALLERY ─────────────────────────────────────────────────────────────────

@app.route('/api/gallery',methods=['GET'])
@login_required
def get_gallery():
    try:
        uid=session['user_id']; role=session.get('role'); conn=get_db()
        rows=conn.execute("SELECT id,user_id,user_name,thought,image_filename,image_type,created_at,updated_at FROM gallery ORDER BY COALESCE(updated_at,created_at) DESC").fetchall()
        conn.close(); return jsonify([serialize_gallery_row(r,uid=uid,role=role) for r in rows])
    except Exception as e:
        logger.error(f"get_gallery error: {e}"); return jsonify({'error':'Failed to load gallery'}),500

@app.route('/api/gallery',methods=['POST'])
@login_required
def post_gallery():
    uid=session['user_id']; name=session.get('user_name','User'); d=request.json or {}
    payload,err=validate_gallery_payload(d,require_image=True)
    if err: return jsonify({'success':False,'message':err[0]}),err[1]
    try:
        decoded=base64.b64decode(payload['image_data'])
        ext=payload['image_type'].split('/')[1] if '/' in payload['image_type'] else 'jpg'
        filename=str(uuid.uuid4())+'.'+ext
        save_photo_file(filename, decoded)          # local + network share
    except Exception as e:
        logger.error(f"save image error: {e}"); return jsonify({'success':False,'message':'Failed to save image'}),500
    now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'); conn=get_db()
    try:
        cur=conn.execute("INSERT INTO gallery (user_id,user_name,thought,image_filename,image_type,created_at,updated_at) VALUES (?,?,?,?,?,?,?)",
                         (uid,name,payload['thought'],filename,payload['image_type'],now,now))
        conn.commit()
        row=conn.execute("SELECT id,user_id,user_name,thought,image_filename,image_type,created_at,updated_at FROM gallery WHERE id=?",(cur.lastrowid,)).fetchone()
        return jsonify({'success':True,'photo':serialize_gallery_row(row,uid=uid,role=session.get('role'))})
    except Exception as e:
        logger.error(f"insert photo error: {e}")
        delete_photo_file(filename)
        return jsonify({'success':False,'message':'Failed to save photo'}),500
    finally: conn.close()

@app.route('/api/gallery/<int:gid>',methods=['PUT'])
@login_required
def update_gallery(gid):
    uid=session['user_id']; now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'); d=request.json or {}
    payload,err=validate_gallery_payload(d,require_image=False)
    if err: return jsonify({'success':False,'message':err[0]}),err[1]
    conn=get_db()
    try:
        row=conn.execute("SELECT * FROM gallery WHERE id=?",(gid,)).fetchone()
        if not row: return jsonify({'success':False,'message':'Photo not found'}),404
        if not is_gallery_manager(row,uid=uid): return jsonify({'success':False,'message':'Not authorized'}),403
        img_fn=row['image_filename']; img_type=row['image_type']
        if payload['image_data']:
            delete_photo_file(row['image_filename'])   # remove old from local + network
            decoded=base64.b64decode(payload['image_data'])
            ext=payload['image_type'].split('/')[1] if '/' in payload['image_type'] else 'jpg'
            img_fn=str(uuid.uuid4())+'.'+ext
            save_photo_file(img_fn, decoded)           # save new to local + network
            img_type=payload['image_type']
        conn.execute("UPDATE gallery SET thought=?,image_filename=?,image_type=?,updated_at=? WHERE id=?",(payload['thought'],img_fn,img_type,now,gid))
        conn.commit()
        updated=conn.execute("SELECT id,user_id,user_name,thought,image_filename,image_type,created_at,updated_at FROM gallery WHERE id=?",(gid,)).fetchone()
        return jsonify({'success':True,'photo':serialize_gallery_row(updated,uid=uid,role=session.get('role'))})
    finally: conn.close()

@app.route('/api/gallery/<int:gid>',methods=['DELETE'])
@login_required
def delete_gallery(gid):
    uid=session['user_id']; conn=get_db()
    try:
        row=conn.execute("SELECT * FROM gallery WHERE id=?",(gid,)).fetchone()
        if not row: return jsonify({'success':False,'message':'Not found'}),404
        if session.get('role')!='admin' and row['user_id']!=uid: return jsonify({'success':False,'message':'Not authorized'}),403
        delete_photo_file(row['image_filename'])       # local + network share
        conn.execute("DELETE FROM gallery WHERE id=?",(gid,)); conn.commit()
        return jsonify({'success':True})
    finally: conn.close()

# ─── USER DASHBOARD ──────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    if session.get('role')=='admin': return redirect(url_for('admin_dash'))
    u=current_user(); cfg=get_db().execute("SELECT * FROM config WHERE id=1").fetchone()
    if u:
        country,country_code,country_flag,display=normalize_country(u['country'],u['country_code'],u['country_flag'])
        u=dict(u); u['country']=country; u['country_code']=country_code; u['country_flag']=country_flag; u['country_display']=display
    return render_template('dashboard.html',user=u,config=cfg)

@app.route('/api/dashboard-data')
@login_required
def dashboard_data():
    try:
        uid=session['user_id']; conn=get_db()
        user=conn.execute("SELECT u.*,t.name as team_name FROM users u LEFT JOIN teams t ON t.id=u.team_id WHERE u.id=?",(uid,)).fetchone()
        cfg=conn.execute("SELECT * FROM config WHERE id=1").fetchone()
        entries=conn.execute("SELECT * FROM steps WHERE user_id=? ORDER BY date ASC",(uid,)).fetchall()
        w=user['weight_kg']; daily_goal=cfg['daily_goal'] if cfg else 10000; event_name=cfg['event_name'] if cfg else 'Walk-a-Thon'; td=today_str()
        uc,ucc,ucf,ucd=normalize_country(user['country'],user['country_code'],user['country_flag'])
        total_steps=sum(e['steps'] for e in entries); today_steps=sum(e['steps'] for e in entries if e['date']==td)
        yest=str(date.today()-timedelta(days=1)); yest_steps=sum(e['steps'] for e in entries if e['date']==yest)
        chart_map={}
        for e in entries: chart_map[e['date']]=chart_map.get(e['date'],0)+e['steps']
        sorted_dates=sorted(chart_map)[-30:]
        streak=calc_streak(uid)
        hist_rows=conn.execute("SELECT * FROM steps WHERE user_id=? ORDER BY date DESC,created_at DESC LIMIT 50",(uid,)).fetchall()
        history=[{'id':e['id'],'date':e['date'],'steps':e['steps'],'km':steps_to_km(e['steps']),'calories':steps_to_cal(e['steps'],w),'note':e['note'] or '','source':e['source']} for e in hist_rows]
        conn.close()
        return jsonify({'user':{'name':user['name'],'team':user['team_name'],'weight_kg':w,'country':uc,'country_flag':ucf,'country_code':ucc,'country_display':ucd},
                        'event_name':event_name,'daily_goal':daily_goal,'total_steps':total_steps,'total_km':steps_to_km(total_steps),
                        'total_calories':steps_to_cal(total_steps,w),'today_steps':today_steps,'today_km':steps_to_km(today_steps),
                        'today_calories':steps_to_cal(today_steps,w),'streak':streak,'badges':get_badges(total_steps,streak),
                        'motivation':motivation(today_steps,daily_goal),'improvement':today_steps-yest_steps,
                        'chart_labels':sorted_dates,'chart_values':[chart_map[d] for d in sorted_dates],'history':history})
    except Exception as e:
        logger.error(f"dashboard_data error: {e}"); return jsonify({'error':'Failed to load dashboard data'}),500

# ─── LOG / EDIT / DELETE STEPS ───────────────────────────────────────────────

@app.route('/api/log-steps',methods=['POST'])
@login_required
def log_steps():
    d=request.json; steps=int(d.get('steps',0)); source=d.get('source','manual')
    note=(d.get('note') or '').strip(); entry_date=d.get('date',today_str())
    try:
        ed=date.fromisoformat(entry_date)
        if ed>date.today(): return jsonify({'success':False,'message':'Cannot log steps for future dates'}),400
    except ValueError: return jsonify({'success':False,'message':'Invalid date'}),400
    if steps<=0: return jsonify({'success':False,'message':'Steps must be greater than 0'}),400
    if steps>100000: return jsonify({'success':False,'message':'Max 100,000 steps per entry'}),400
    uid=session['user_id']; conn=get_db()
    existing=conn.execute("SELECT SUM(steps) as total FROM steps WHERE user_id=? AND date=?",(uid,entry_date)).fetchone()['total'] or 0
    if existing+steps>100000: conn.close(); return jsonify({'success':False,'message':'Daily step limit exceeded'}),400
    now=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        conn.execute("INSERT INTO steps (user_id,steps,date,note,source,created_at,updated_at) VALUES (?,?,?,?,?,?,?)",(uid,steps,entry_date,note or None,source,now,now))
        conn.commit()
        total=conn.execute("SELECT SUM(steps) as t FROM steps WHERE user_id=?",(uid,)).fetchone()['t'] or 0
        w=conn.execute("SELECT weight_kg FROM users WHERE id=?",(uid,)).fetchone()['weight_kg']
        prev=total-steps; milestone=None
        if prev<10000<=total: milestone='🎉 You crossed 10,000 total steps!'
        elif prev<50000<=total: milestone='🔥 Epic! 50,000 total steps!'
        elif prev<100000<=total: milestone='🏆 LEGENDARY! 100,000 steps!'
        return jsonify({'success':True,'milestone':milestone,'calories_burned':steps_to_cal(steps,w)})
    finally: conn.close()

@app.route('/api/live-step-session',methods=['POST'])
@login_required
def live_step_session():
    uid=session['user_id']; d=request.json or {}
    session_id=(d.get('session_id') or '').strip(); entry_date=(d.get('date') or today_str()).strip()
    mode=(d.get('mode') or 'auto').strip().lower(); status=(d.get('status') or 'running').strip().lower()
    try: steps=int(d.get('steps',0))
    except: return jsonify({'success':False,'message':'Steps must be a number'}),400
    if not session_id or len(session_id)>64: return jsonify({'success':False,'message':'Valid session ID required'}),400
    if mode not in {'auto','manual'}: mode='auto'
    if status not in {'running','stopped'}: status='running'
    if steps<0: return jsonify({'success':False,'message':'Steps cannot be negative'}),400
    if steps>100000: return jsonify({'success':False,'message':'Max 100,000 steps per session'}),400
    try:
        dt=datetime.strptime(entry_date,'%Y-%m-%d').date()
        if dt>date.today(): return jsonify({'success':False,'message':'Cannot log steps for future dates'}),400
    except: return jsonify({'success':False,'message':'Invalid date format'}),400
    now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'); note=f'Hybrid {mode} session ({status})'
    conn=get_db()
    try:
        row=conn.execute("SELECT id FROM steps WHERE user_id=? AND date=? AND session_id=?",(uid,entry_date,session_id)).fetchone()
        if steps==0 and not row: return jsonify({'success':True,'created':False,'entry_id':None,'session_id':session_id,'steps':0})
        if row:
            conn.execute("UPDATE steps SET steps=?,note=?,source=?,updated_at=? WHERE id=?",(steps,note,mode,now,row['id'])); entry_id=row['id']; created=False
        else:
            cur=conn.execute("INSERT INTO steps (user_id,steps,date,note,source,session_id,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?)",(uid,steps,entry_date,note,mode,session_id,now,now)); entry_id=cur.lastrowid; created=True
        conn.commit()
        return jsonify({'success':True,'created':created,'entry_id':entry_id,'session_id':session_id,'steps':steps})
    finally: conn.close()

@app.route('/api/steps/<int:sid>',methods=['PUT'])
@login_required
def edit_step(sid):
    d=request.json; steps=int(d.get('steps',0)); note=(d.get('note') or '').strip(); entry_date=d.get('date',today_str())
    try:
        ed=date.fromisoformat(entry_date)
        if ed>date.today(): return jsonify({'success':False,'message':'Cannot set future date'}),400
    except: return jsonify({'success':False,'message':'Invalid date'}),400
    if steps<=0 or steps>100000: return jsonify({'success':False,'message':'Steps must be 1–100,000'}),400
    uid=session['user_id']; now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'); conn=get_db()
    try:
        row=conn.execute("SELECT * FROM steps WHERE id=?",(sid,)).fetchone()
        if not row: return jsonify({'success':False,'message':'Entry not found'}),404
        if session.get('role')!='admin' and row['user_id']!=uid: return jsonify({'success':False,'message':'Not authorized'}),403
        conn.execute("UPDATE steps SET steps=?,date=?,note=?,updated_at=? WHERE id=?",(steps,entry_date,note or None,now,sid)); conn.commit()
        return jsonify({'success':True})
    finally: conn.close()

@app.route('/api/steps/<int:sid>',methods=['DELETE'])
@login_required
def delete_step(sid):
    uid=session['user_id']; conn=get_db()
    try:
        row=conn.execute("SELECT * FROM steps WHERE id=?",(sid,)).fetchone()
        if not row: return jsonify({'success':False,'message':'Entry not found'}),404
        if session.get('role')!='admin' and row['user_id']!=uid: return jsonify({'success':False,'message':'Not authorized'}),403
        conn.execute("DELETE FROM steps WHERE id=?",(sid,)); conn.commit(); return jsonify({'success':True})
    finally: conn.close()

# ─── PROFILE ─────────────────────────────────────────────────────────────────

@app.route('/api/profile',methods=['GET','POST'])
@login_required
def profile():
    uid=session['user_id']; conn=get_db()
    try:
        if request.method=='POST':
            d=request.json; weight=d.get('weight_kg')
            weight_kg=float(weight) if weight and str(weight).strip() else None
            conn.execute("UPDATE users SET weight_kg=? WHERE id=?",(weight_kg,uid)); conn.commit()
            return jsonify({'success':True})
        u=conn.execute("SELECT name,email,weight_kg FROM users WHERE id=?",(uid,)).fetchone()
        return jsonify(dict(u))
    finally: conn.close()

# ─── TEAM ────────────────────────────────────────────────────────────────────

@app.route('/team')
@login_required
def team_page():
    if session.get('role')=='admin': return redirect(url_for('admin_dash'))
    return render_template('team.html',user=current_user())

@app.route('/api/team-data')
@login_required
def team_data():
    uid=session['user_id']; conn=get_db()
    user=conn.execute("SELECT * FROM users WHERE id=?",(uid,)).fetchone()
    if not user['team_id']: conn.close(); return jsonify({'in_team':False})
    members=conn.execute("SELECT u.id,u.name,u.country,u.country_code,u.country_flag FROM users u WHERE u.team_id=? ORDER BY u.name",(user['team_id'],)).fetchall()
    team=conn.execute("SELECT * FROM teams WHERE id=?",(user['team_id'],)).fetchone()
    team_steps=conn.execute("SELECT SUM(s.steps) as t FROM steps s JOIN users u ON u.id=s.user_id WHERE u.team_id=?",(user['team_id'],)).fetchone()['t'] or 0
    today_team=conn.execute("SELECT SUM(s.steps) as t FROM steps s JOIN users u ON u.id=s.user_id WHERE u.team_id=? AND s.date=?",(user['team_id'],today_str())).fetchone()['t'] or 0
    conn.close()
    ml=[]
    for m in members:
        cnt,cc,cf,cd=normalize_country(m['country'],m['country_code'],m['country_flag'])
        ml.append({'id':m['id'],'name':m['name'],'is_me':m['id']==uid,'country':cnt,'country_flag':cf,'country_display':cd})
    return jsonify({'in_team':True,'team_name':team['name'],'invite_code':team['invite_code'],'members':ml,'team_total':team_steps,'team_km':steps_to_km(team_steps),'today_total':today_team})

@app.route('/api/join-team',methods=['POST'])
@login_required
def join_team():
    uid=session['user_id']; d=request.json or {}; invite_code=(d.get('invite_code') or '').strip().upper()
    if not invite_code: return jsonify({'success':False,'message':'Invite code is required'}),400
    conn=get_db()
    try:
        user=conn.execute("SELECT team_id FROM users WHERE id=?",(uid,)).fetchone()
        if user['team_id']: return jsonify({'success':False,'message':'You are already in a team'}),400
        team=conn.execute("SELECT id FROM teams WHERE invite_code=?",(invite_code,)).fetchone()
        if not team: return jsonify({'success':False,'message':'Invalid invite code'}),400
        conn.execute("UPDATE users SET team_id=? WHERE id=?",(team['id'],uid)); conn.commit()
        return jsonify({'success':True})
    finally: conn.close()

@app.route('/api/leave-team',methods=['POST'])
@login_required
def leave_team():
    uid=session['user_id']; conn=get_db()
    try:
        conn.execute("UPDATE users SET team_id=NULL WHERE id=?",(uid,)); conn.commit()
        return jsonify({'success':True})
    finally: conn.close()

# ─── LEADERBOARD ─────────────────────────────────────────────────────────────

@app.route('/leaderboard')
@app.route('/stepcounter')
@login_required
def leaderboard_page():
    if session.get('role')=='admin': return redirect(url_for('admin_dash'))
    return render_template('stepcounter.html',user=current_user())

@app.route('/api/leaderboard')
@login_required
def leaderboard():
    conn=get_db()
    rows=conn.execute('''SELECT u.id,u.name,u.weight_kg,u.country,u.country_code,u.country_flag,t.name as team_name,COALESCE(SUM(s.steps),0) as total_steps
        FROM users u LEFT JOIN teams t ON t.id=u.team_id LEFT JOIN steps s ON s.user_id=u.id WHERE u.role!='admin' GROUP BY u.id ORDER BY total_steps DESC''').fetchall()
    individual=[]
    for i,r in enumerate(rows):
        cnt,cc,cf,cd=normalize_country(r['country'],r['country_code'],r['country_flag'])
        individual.append({'rank':i+1,'name':r['name'],'team':r['team_name'] or '—','country':cnt,'country_flag':cf,'country_display':cd,'total_steps':r['total_steps'],'total_km':steps_to_km(r['total_steps']),'total_calories':steps_to_cal(r['total_steps'],r['weight_kg'])})
    teams_rows=conn.execute('''SELECT t.name,COUNT(DISTINCT u.id) as members,COALESCE(SUM(s.steps),0) as total_steps
        FROM teams t JOIN users u ON u.team_id=t.id LEFT JOIN steps s ON s.user_id=u.id GROUP BY t.id ORDER BY total_steps DESC''').fetchall()
    teams=[{'rank':i+1,'team':r['name'],'members':r['members'],'total_steps':r['total_steps'],'total_km':steps_to_km(r['total_steps'])} for i,r in enumerate(teams_rows)]
    today_rows=conn.execute('''SELECT u.name,t.name as team_name,SUM(s.steps) as steps FROM steps s JOIN users u ON u.id=s.user_id LEFT JOIN teams t ON t.id=u.team_id WHERE s.date=? GROUP BY u.id ORDER BY steps DESC LIMIT 10''',(today_str(),)).fetchall()
    conn.close()
    return jsonify({'individual':individual,'teams':teams,'today_active':[dict(r) for r in today_rows]})

# ─── FITBIT ───────────────────────────────────────────────────────────────────

@app.route('/fitbit')
@login_required
def fitbit_page():
    if session.get('role')=='admin': return redirect(url_for('admin_dash'))
    return render_template('fitbit.html',user=current_user())

@app.route('/fitbit/connect')
@login_required
def fitbit_connect():
    conn=get_db(); cfg=conn.execute("SELECT fitbit_client_id FROM config WHERE id=1").fetchone(); conn.close()
    client_id=cfg['fitbit_client_id'] if cfg else ''
    if not client_id: return "Fitbit not configured by admin.",400
    from urllib.parse import urlencode
    params={'response_type':'code','client_id':client_id,'redirect_uri':get_fitbit_redirect_uri(),'scope':'activity','expires_in':'604800'}
    return redirect(f"{FITBIT_AUTH_URL}?{urlencode(params)}")

def get_fitbit_redirect_uri():
    if FITBIT_REDIRECT: return FITBIT_REDIRECT.rstrip('/')
    return url_for('fitbit_callback',_external=True,_scheme='https')

@app.route('/fitbit/callback')
@login_required
def fitbit_callback():
    code=request.args.get('code')
    if not code: return "Fitbit auth failed.",400
    conn=get_db(); cfg=conn.execute("SELECT fitbit_client_id,fitbit_client_secret FROM config WHERE id=1").fetchone()
    cid=cfg['fitbit_client_id'] if cfg else ''; cs=cfg['fitbit_client_secret'] if cfg else ''
    creds=base64.b64encode(f"{cid}:{cs}".encode()).decode()
    resp=req_lib.post(FITBIT_TOKEN_URL,headers={'Authorization':f'Basic {creds}','Content-Type':'application/x-www-form-urlencoded'},
                      data={'code':code,'grant_type':'authorization_code','redirect_uri':get_fitbit_redirect_uri()})
    if resp.status_code!=200: conn.close(); return f"Token exchange failed: {resp.text}",400
    tokens=resp.json()
    conn.execute("UPDATE users SET fitbit_access_token=?,fitbit_refresh_token=?,fitbit_user_id=? WHERE id=?",(tokens['access_token'],tokens.get('refresh_token',''),tokens.get('user_id',''),session['user_id']))
    conn.commit(); conn.close()
    return redirect(url_for('fitbit_page'))

@app.route('/api/fitbit/sync',methods=['POST'])
@login_required
def fitbit_sync():
    uid=session['user_id']; conn=get_db()
    user=conn.execute("SELECT fitbit_access_token,fitbit_refresh_token FROM users WHERE id=?",(uid,)).fetchone()
    if not user or not user['fitbit_access_token']: conn.close(); return jsonify({'success':False,'message':'Fitbit not connected'}),400
    resp=req_lib.get(FITBIT_STEPS_URL.format(date=today_str()),headers={'Authorization':f"Bearer {user['fitbit_access_token']}"})
    if resp.status_code==401: conn.close(); return jsonify({'success':False,'message':'Fitbit token expired.'}),401
    if resp.status_code!=200: conn.close(); return jsonify({'success':False,'message':'Fitbit API error'}),500
    data=resp.json(); steps=int(data.get('activities-steps',[{}])[0].get('value',0))
    now=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute("INSERT INTO steps (user_id,steps,date,source,created_at,updated_at) VALUES (?,?,?,?,?,?)",(uid,steps,today_str(),'fitbit',now,now))
    conn.commit(); conn.close()
    return jsonify({'success':True,'steps':steps})

@app.route('/api/fitbit/disconnect',methods=['POST'])
@login_required
def fitbit_disconnect():
    uid=session['user_id']; conn=get_db()
    conn.execute("UPDATE users SET fitbit_access_token=NULL,fitbit_refresh_token=NULL,fitbit_user_id=NULL WHERE id=?",(uid,))
    conn.commit(); conn.close(); return jsonify({'success':True})

# ─── NOTICES ─────────────────────────────────────────────────────────────────

@app.route('/api/notices')
@login_required
def get_notices():
    conn=get_db(); rows=conn.execute("SELECT * FROM notices WHERE is_active=1 ORDER BY created_at DESC LIMIT 20").fetchall(); conn.close()
    return jsonify([dict(r) for r in rows])

# ─── ADMIN PAGES ─────────────────────────────────────────────────────────────

@app.route('/admin')
@admin_required
def admin_dash():
    cfg=get_db().execute("SELECT * FROM config WHERE id=1").fetchone()
    return render_template('admin/dashboard.html',config=cfg,user_name=session.get('user_name'))

@app.route('/admin/users')
@admin_required
def admin_users():
    conn=get_db()
    users=conn.execute('''SELECT u.id,u.name,u.email,u.role,u.weight_kg,u.country,u.country_flag,u.created_at,t.name as team_name,COALESCE(SUM(s.steps),0) as total_steps
        FROM users u LEFT JOIN teams t ON t.id=u.team_id LEFT JOIN steps s ON s.user_id=u.id GROUP BY u.id ORDER BY total_steps DESC''').fetchall()
    conn.close(); return render_template('admin/users.html',users=users,user_name=session.get('user_name'))

@app.route('/admin/leaderboard')
@admin_required
def admin_leaderboard():
    return render_template('admin/leaderboard.html',user_name=session.get('user_name'))

@app.route('/admin/steps')
@admin_required
def admin_steps():
    return render_template('admin/steps.html',user_name=session.get('user_name'))

@app.route('/admin/photos')
@admin_required
def admin_photos():
    return render_template('admin/photos.html',user_name=session.get('user_name'))

# ─── ADMIN API ───────────────────────────────────────────────────────────────

@app.route('/api/admin/stats')
@admin_required
def admin_stats():
    conn=get_db()
    tu=conn.execute("SELECT COUNT(*) as c FROM users WHERE role!='admin'").fetchone()['c']
    tt=conn.execute("SELECT COUNT(*) as c FROM teams WHERE EXISTS (SELECT 1 FROM users WHERE users.team_id=teams.id)").fetchone()['c']
    ts=conn.execute("SELECT COALESCE(SUM(steps),0) as t FROM steps").fetchone()['t']
    ta=conn.execute("SELECT COUNT(DISTINCT user_id) as c FROM steps WHERE date=?",(today_str(),)).fetchone()['c']
    conn.close()
    net_ok=False
    try: net_ok=os.path.isdir(_network_excel_dir()) if _network_excel_dir() else False
    except: pass
    return jsonify({'total_users':tu,'total_teams':tt,'total_steps':ts,'total_km':steps_to_km(ts),'today_active':ta,'network_share_ok':net_ok})

@app.route('/api/admin/config',methods=['GET','POST'])
@admin_required
def admin_config():
    conn=get_db()
    try:
        if request.method=='POST':
            d=request.json
            conn.execute("UPDATE config SET event_name=?,daily_goal=?,fitbit_client_id=?,fitbit_client_secret=?,start_date=?,end_date=? WHERE id=1",
                         (d.get('event_name','Walk-a-Thon'),int(d.get('daily_goal',10000)),d.get('fitbit_client_id',''),d.get('fitbit_client_secret',''),d.get('start_date',''),d.get('end_date','')))
            conn.commit(); return jsonify({'success':True})
        row=conn.execute("SELECT * FROM config WHERE id=1").fetchone()
        return jsonify(dict(row))
    finally: conn.close()

@app.route('/api/admin/users/<int:uid>',methods=['DELETE'])
@admin_required
def admin_delete_user(uid):
    conn=get_db()
    try:
        # Delete photos from local + network
        photos=conn.execute("SELECT image_filename FROM gallery WHERE user_id=?",(uid,)).fetchall()
        for p in photos: delete_photo_file(p['image_filename'])
        conn.execute("DELETE FROM gallery WHERE user_id=?",(uid,))
        conn.execute("DELETE FROM steps WHERE user_id=?",(uid,))
        conn.execute("DELETE FROM users WHERE id=?",(uid,))
        conn.execute("DELETE FROM teams WHERE id NOT IN (SELECT DISTINCT team_id FROM users WHERE team_id IS NOT NULL)")
        conn.commit(); return jsonify({'success':True})
    finally: conn.close()

@app.route('/api/admin/notices',methods=['GET'])
@admin_required
def admin_get_notices():
    conn=get_db(); rows=conn.execute("SELECT * FROM notices ORDER BY created_at DESC").fetchall(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/admin/notices',methods=['POST'])
@admin_required
def admin_post_notice():
    d=request.json; msg=(d.get('message') or '').strip()
    if not msg: return jsonify({'success':False,'message':'Message cannot be empty'}),400
    conn=get_db(); conn.execute("INSERT INTO notices (message,created_by) VALUES (?,?)",(msg,session.get('user_name','Admin'))); conn.commit(); conn.close()
    return jsonify({'success':True})

@app.route('/api/admin/notices/<int:nid>',methods=['DELETE'])
@admin_required
def admin_delete_notice(nid):
    conn=get_db(); conn.execute("DELETE FROM notices WHERE id=?",(nid,)); conn.commit(); conn.close()
    return jsonify({'success':True})

@app.route('/api/admin/leaderboard-data')
@admin_required
def admin_lb_data():
    conn=get_db()
    rows=conn.execute('''SELECT u.id,u.name,u.weight_kg,u.country,u.country_code,u.country_flag,t.name as team_name,COALESCE(SUM(s.steps),0) as total_steps
        FROM users u LEFT JOIN teams t ON t.id=u.team_id LEFT JOIN steps s ON s.user_id=u.id WHERE u.role!='admin' GROUP BY u.id ORDER BY total_steps DESC''').fetchall()
    individual=[]
    for i,r in enumerate(rows):
        cnt,cc,cf,cd=normalize_country(r['country'],r['country_code'],r['country_flag'])
        individual.append({'rank':i+1,'name':r['name'],'team':r['team_name'] or '—','country':cnt,'country_flag':cf,'country_display':cd,'total_steps':r['total_steps'],'total_km':steps_to_km(r['total_steps']),'total_calories':steps_to_cal(r['total_steps'],r['weight_kg'])})
    teams_rows=conn.execute('''SELECT t.name,COUNT(DISTINCT u.id) as members,COALESCE(SUM(s.steps),0) as total_steps
        FROM teams t JOIN users u ON u.team_id=t.id LEFT JOIN steps s ON s.user_id=u.id GROUP BY t.id ORDER BY total_steps DESC''').fetchall()
    teams=[{'rank':i+1,'team':r['name'],'members':r['members'],'total_steps':r['total_steps'],'total_km':steps_to_km(r['total_steps'])} for i,r in enumerate(teams_rows)]
    today_rows=conn.execute('''SELECT u.name,t.name as team_name,SUM(s.steps) as steps FROM steps s JOIN users u ON u.id=s.user_id LEFT JOIN teams t ON t.id=u.team_id WHERE s.date=? GROUP BY u.id ORDER BY steps DESC LIMIT 10''',(today_str(),)).fetchall()
    conn.close()
    return jsonify({'individual':individual,'teams':teams,'today_active':[dict(r) for r in today_rows]})

@app.route('/api/admin/steps')
@admin_required
def admin_steps_data():
    page=int(request.args.get('page',1)); per_page=30
    uid_f=request.args.get('user_id'); date_f=request.args.get('date'); conn=get_db()
    where=[]; params=[]
    if uid_f: where.append("s.user_id=?"); params.append(uid_f)
    if date_f: where.append("s.date=?"); params.append(date_f)
    w=('WHERE '+' AND '.join(where)) if where else ''
    total=conn.execute(f"SELECT COUNT(*) as c FROM steps s {w}",params).fetchone()['c']
    rows=conn.execute(f"SELECT s.*,u.name as user_name,u.weight_kg,u.country,u.country_code,u.country_flag FROM steps s JOIN users u ON u.id=s.user_id {w} ORDER BY s.date DESC,s.created_at DESC LIMIT ? OFFSET ?",params+[per_page,(page-1)*per_page]).fetchall()
    users_list=conn.execute("SELECT id,name FROM users WHERE role!='admin' ORDER BY name").fetchall(); conn.close()
    entries=[]
    for r in rows:
        cnt,cc,cf,cd=normalize_country(r['country'],r['country_code'],r['country_flag'])
        entries.append({'id':r['id'],'user_name':r['user_name'],'user_id':r['user_id'],'steps':r['steps'],'date':r['date'],'note':r['note'] or '','source':r['source'],'km':steps_to_km(r['steps']),'calories':steps_to_cal(r['steps'],r['weight_kg']),'created_at':r['created_at'],'country':cnt,'country_flag':cf,'country_display':cd})
    return jsonify({'entries':entries,'total':total,'page':page,'per_page':per_page,'users':[dict(u) for u in users_list]})

# CSV export (unchanged)
@app.route('/api/admin/export')
@admin_required
def admin_export():
    fmt=request.args.get('type','individual'); conn=get_db(); out=io.StringIO(); w=csv.writer(out)
    if fmt=='individual':
        rows=conn.execute('''SELECT u.name,u.email,t.name as team,u.weight_kg,COALESCE(SUM(s.steps),0) as total_steps FROM users u LEFT JOIN teams t ON t.id=u.team_id LEFT JOIN steps s ON s.user_id=u.id WHERE u.role!='admin' GROUP BY u.id ORDER BY total_steps DESC''').fetchall()
        w.writerow(['Rank','Name','Email','Team','Weight(kg)','Total Steps','Total KM','Calories'])
        for i,r in enumerate(rows):
            cal=steps_to_cal(r['total_steps'],r['weight_kg'])
            w.writerow([i+1,r['name'],r['email'],r['team'] or '—',r['weight_kg'] or '—',r['total_steps'],steps_to_km(r['total_steps']),cal or '—'])
    elif fmt=='full':
        rows=conn.execute('''SELECT s.*,u.name as user_name,u.weight_kg,t.name as team_name FROM steps s JOIN users u ON u.id=s.user_id LEFT JOIN teams t ON t.id=u.team_id ORDER BY s.date DESC''').fetchall()
        w.writerow(['User','Team','Steps','KM','Calories','Date','Note','Source','Logged At'])
        for r in rows:
            w.writerow([r['user_name'],r['team_name'] or '—',r['steps'],steps_to_km(r['steps']),steps_to_cal(r['steps'],r['weight_kg']) or '—',r['date'],r['note'] or '',r['source'],r['created_at']])
    conn.close(); out.seek(0)
    return Response(out.getvalue(),mimetype='text/csv',headers={'Content-Disposition':f'attachment;filename={fmt}.csv'})

# Excel export (on-demand download)
@app.route('/api/admin/export-excel')
@admin_required
def admin_export_excel():
    try:
        wb=build_excel_workbook(); buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        ts=datetime.now().strftime('%Y%m%d_%H%M%S')
        return Response(buf.getvalue(),
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition':f'attachment;filename=walkathon_backup_{ts}.xlsx'})
    except Exception as e:
        logger.error(f'Excel export error: {e}')
        return jsonify({'success':False,'message':str(e)}),500

# Trigger immediate network backup
@app.route('/api/admin/trigger-backup',methods=['POST'])
@admin_required
def admin_trigger_backup():
    path=export_excel_to_network()
    if path: return jsonify({'success':True,'path':path})
    return jsonify({'success':False,'message':'Backup failed – check server logs'}),500

# Test network share connectivity
@app.route('/api/admin/test-network',methods=['GET'])
@admin_required
def admin_test_network():
    ok = connect_network_share()
    if ok:
        _ensure_network_dirs()
    excel_ok = os.path.isdir(_network_excel_dir()) if _network_excel_dir() else False
    photos_ok = os.path.isdir(_network_photos_dir()) if _network_photos_dir() else False
    od_excel_ok  = os.path.isdir(ONEDRIVE_EXCEL_DIR)  if ONEDRIVE_EXCEL_DIR  else False
    od_photos_ok = os.path.isdir(ONEDRIVE_PHOTOS_DIR) if ONEDRIVE_PHOTOS_DIR else False
    return jsonify({
        'connected': ok,
        'share': NETWORK_SHARE,
        'excel_dir': _network_excel_dir(),
        'photos_dir': _network_photos_dir(),
        'excel_dir_ok': excel_ok,
        'photos_dir_ok': photos_ok,
        'onedrive_path': ONEDRIVE_BACKUP_PATH,
        'onedrive_excel_ok': od_excel_ok,
        'onedrive_photos_ok': od_photos_ok,
    })

# Upload Excel backup to restore DB
@app.route('/api/admin/upload-backup',methods=['POST'])
@admin_required
def admin_upload_backup():
    if 'file' not in request.files: return jsonify({'success':False,'message':'No file uploaded'}),400
    f=request.files['file']
    if not f.filename: return jsonify({'success':False,'message':'Empty filename'}),400
    if not f.filename.lower().endswith('.xlsx'): return jsonify({'success':False,'message':'Only .xlsx files accepted'}),400
    count,err=restore_from_excel(f.stream)
    if err: return jsonify({'success':False,'message':err}),400
    return jsonify({'success':True,'message':f'Restore complete. {count} records imported successfully.'})

# Admin photos
@app.route('/api/admin/photos')
@admin_required
def admin_photos_data():
    conn=get_db(); rows=conn.execute("SELECT id,user_id,user_name,thought,image_filename,image_type,created_at,updated_at FROM gallery ORDER BY COALESCE(updated_at,created_at) DESC").fetchall(); conn.close()
    uid=session['user_id']; role=session.get('role')
    return jsonify([serialize_gallery_row(r,uid=uid,role=role) for r in rows])

# ─── SERVE UPLOADS ────────────────────────────────────────────────────────────

@app.route('/uploads/<filename>')
@app.route('/static/uploads/<filename>')
def serve_upload(filename):
    """
    Serve a photo file.
    If the file is missing locally (e.g. after a server crash / migration),
    it is automatically restored from the network share before serving.
    """
    local_path = os.path.join(LOCAL_UPLOADS_DIR, filename)
    if not os.path.exists(local_path):
        restore_photo_from_network(filename)   # try to pull from share
    try:
        return send_from_directory(LOCAL_UPLOADS_DIR, filename, max_age=86400)
    except Exception:
        return 'File not found', 404

# ─── MAIN ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()

    # Connect to \\10.211.243.115\walkdata and create sub-folders
    logger.info(f'🔗 Connecting to network share {NETWORK_SHARE} as {NETWORK_USER} …')
    if connect_network_share():
        _ensure_network_dirs()
    else:
        logger.warning('⚠️  Network share unavailable at startup – will use local fallback. Retried on each backup.')

    start_backup_scheduler()   # hourly Excel backup → network + local

    host=os.environ.get('HOST','0.0.0.0'); port=int(os.environ.get('PORT','443'))
    debug=env_flag('FLASK_DEBUG',False); ssl_context=get_ssl_context()
    scheme='https' if ssl_context else 'http'; local_host='localhost' if host in {'0.0.0.0','::'} else host
    logger.info(f'🚀 Server: {scheme}://{local_host}:{port}')
    logger.info(f'   Share  : {NETWORK_SHARE}/excel + /photos')
    logger.info(f'   OneDrive: {ONEDRIVE_BACKUP_PATH}')
    logger.info(f'   DB    : {DB_PATH}')
    logger.info(f'   Admin : admin@walkathon.com / Flask@Mitel#Walkathon26!')
    app.run(host=host,port=port,debug=debug,ssl_context=ssl_context)
